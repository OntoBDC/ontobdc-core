from __future__ import annotations

import hashlib
import json
import re
from csv import DictReader
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping
from urllib.parse import quote

from rdflib import Graph, Namespace, URIRef
from rdflib.compare import to_canonical_graph
from rdflib.namespace import DCTERMS, RDF

from ontobdc.cli.domain.port.context import CliContextPort
from ontobdc.shared.adapter.capability import TransactionCapability
from ontobdc.shared.domain.model.capability import CapabilityMetadata
from ontobdc.view.domain.machine.state import ContainerViewProcessState
from ontobdc.view.plugin.capability.hardcoded import HardcodedCapability


OBDC = Namespace("http://ontobdc.org/ontology/domain/ns.ttl#")
SCHEMA = Namespace("https://schema.org/")
METADATA_DIRECTORY = ".__ontobdc__"
DATASET_METADATA_FILENAME = "dataset.ttl"
DATAPACKAGE_FILENAME = "datapackage.json"
FACADE_RELATIVE_PATH = Path(METADATA_DIRECTORY) / "linkset" / "facade.ttl"
FACADES_JSONLD_RELATIVE_PATH = (
    Path(METADATA_DIRECTORY) / "view" / "facades.jsonld"
)
INDEX_FILENAME = "index.html"
SCRIPT_ID = "ontobdc-dataset-facades"
_FINGERPRINT_ATTRIBUTE = "data-ontobdc-fingerprint"
_SCRIPT_PATTERN = re.compile(
    rf"[ \t]*<script\b(?=[^>]*\bid=[\"']{re.escape(SCRIPT_ID)}[\"'])"
    r"[^>]*>.*?</script>\s*",
    re.IGNORECASE | re.DOTALL,
)

HAS_FACADE_FIELD_LOCAL_NAME = "hasFacadeField"
FIELD_DATATYPE_LOCAL_NAME = "fieldDatatype"
MAPS_TO_PROPERTY_LOCAL_NAME = "mapsToProperty"
MAPS_FROM_HEADER_FIELD_LOCAL_NAME = "mapsFromHeaderField"
IFC_CLASS_NAME_LOCAL_NAME = "ifcClassName"
INSTANCE_ROLE_LOCAL_NAME = "instanceRole"
IS_FACADE_OF_LOCAL_NAME = "isFacadeOf"


def resolve_container_path(context: CliContextPort) -> Path:
    raw_value = context.get_parameter_value("container_path")
    normalized = str(raw_value or "").strip()
    if not normalized:
        raise ValueError("The container path was not resolved.")
    container_path = Path(normalized).expanduser().resolve()
    if not container_path.is_dir():
        raise ValueError(
            f"Container path is not an accessible directory: {container_path}"
        )
    return container_path


def dataset_facade_files(container_path: Path) -> tuple[Path, ...]:
    """Return facade files for the semantic datasets directly in a container."""
    facade_files = []
    for dataset_path in sorted(
        (path for path in container_path.iterdir() if path.is_dir()),
        key=lambda path: path.name,
    ):
        dataset_metadata = (
            dataset_path / METADATA_DIRECTORY / DATASET_METADATA_FILENAME
        )
        if not dataset_metadata.is_file():
            continue

        metadata_graph = Graph()
        try:
            metadata_graph.parse(dataset_metadata, format="turtle")
        except Exception as exc:
            raise ValueError(
                f"Invalid dataset metadata: {dataset_metadata}"
            ) from exc

        if not any(
            True
            for _ in metadata_graph.subjects(RDF.type, OBDC.EntityDataset)
        ):
            continue

        facade_file = dataset_path / FACADE_RELATIVE_PATH
        if not facade_file.is_file():
            raise ValueError(
                "Dataset facade not found: "
                f"{facade_file.relative_to(container_path)}"
            )
        facade_files.append(facade_file)

    return tuple(facade_files)


def calculate_facades_fingerprint(
    container_path: Path,
    facade_files: Iterable[Path] | None = None,
) -> str:
    selected_files = tuple(
        facade_files
        if facade_files is not None
        else dataset_serialization_files(container_path)
    )
    digest = hashlib.sha256()
    for source_file in selected_files:
        relative_path = source_file.relative_to(container_path).as_posix()
        digest.update(relative_path.encode("utf-8"))
        digest.update(b"\0")
        digest.update(source_file.read_bytes())
        digest.update(b"\0")
    return digest.hexdigest()


def dataset_serialization_files(container_path: Path) -> tuple[Path, ...]:
    source_files: set[Path] = set()
    for facade_file in dataset_facade_files(container_path):
        dataset_path = facade_file.parents[2]
        metadata_path = (
            dataset_path / METADATA_DIRECTORY / DATASET_METADATA_FILENAME
        )
        datapackage_path = (
            dataset_path / METADATA_DIRECTORY / DATAPACKAGE_FILENAME
        )
        if not datapackage_path.is_file():
            raise ValueError(
                "Dataset datapackage not found: "
                f"{datapackage_path.relative_to(container_path)}"
            )
        descriptor = _load_json_object(datapackage_path, "dataset datapackage")
        source_files.update(
            {metadata_path, datapackage_path, facade_file}
        )
        for resource in _resource_descriptors(descriptor):
            source_files.add(
                _resource_path(dataset_path, datapackage_path, resource)
            )
    return tuple(
        sorted(
            source_files,
            key=lambda path: path.relative_to(container_path).as_posix(),
        )
    )


def _load_json_object(path: Path, label: str) -> Dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Invalid {label}: {path}") from exc
    if not isinstance(value, dict):
        raise ValueError(f"Invalid {label}: {path}")
    return value


def _resource_descriptors(descriptor: Mapping[str, Any]) -> list[Dict[str, Any]]:
    resources = descriptor.get("resources", [])
    if not isinstance(resources, list):
        raise ValueError("Dataset datapackage resources must be a list.")
    return [item for item in resources if isinstance(item, dict)]


def _resource_path(
    dataset_path: Path,
    datapackage_path: Path,
    resource: Mapping[str, Any],
) -> Path:
    raw_path = str(resource.get("path") or "").strip()
    if not raw_path:
        raise ValueError(
            f"Dataset resource path is required: {datapackage_path}"
        )
    candidate = Path(raw_path).expanduser()
    resolved = (
        candidate.resolve()
        if candidate.is_absolute()
        else (datapackage_path.parent / candidate).resolve()
    )
    try:
        resolved.relative_to(dataset_path.resolve())
    except ValueError as exc:
        raise ValueError(
            f"Dataset resource escapes its dataset directory: {raw_path}"
        ) from exc
    if not resolved.is_file():
        raise ValueError(f"Dataset resource not found: {resolved}")
    return resolved


def _local_name(value: Any) -> str:
    text = str(value)
    return text.rsplit("#", 1)[-1].rsplit("/", 1)[-1]


def _predicate_by_local_name(
    graph: Graph,
    subject: Any,
    local_name: str,
) -> Any | None:
    return next(
        (
            value
            for predicate, value in graph.predicate_objects(subject)
            if _local_name(predicate) == local_name
        ),
        None,
    )


def _facade_contract(
    graph: Graph,
    facade_uri: str,
) -> Dict[str, Any]:
    facade = URIRef(facade_uri)
    if not any(graph.triples((facade, None, None))):
        raise ValueError(f"Facade is not declared by facade.ttl: {facade_uri}")

    fields: Dict[str, Dict[str, str]] = {}
    for predicate, field in graph.predicate_objects(facade):
        if _local_name(predicate) != HAS_FACADE_FIELD_LOCAL_NAME:
            continue
        identifier = str(graph.value(field, SCHEMA.identifier) or "").strip()
        if not identifier:
            raise ValueError(
                f"Facade field does not declare schema:identifier: {field}"
            )
        mapped_predicate = (
            _predicate_by_local_name(graph, field, MAPS_TO_PROPERTY_LOCAL_NAME)
            or _predicate_by_local_name(
                graph,
                field,
                MAPS_FROM_HEADER_FIELD_LOCAL_NAME,
            )
        )
        if mapped_predicate is None:
            raise ValueError(
                f"Facade field does not map to a property: {field}"
            )
        datatype = _predicate_by_local_name(
            graph,
            field,
            FIELD_DATATYPE_LOCAL_NAME,
        )
        fields[identifier] = {
            "predicate": str(mapped_predicate),
            "datatype": str(datatype or ""),
        }

    class_name = _predicate_by_local_name(
        graph,
        facade,
        IFC_CLASS_NAME_LOCAL_NAME,
    )
    role = _predicate_by_local_name(
        graph,
        facade,
        INSTANCE_ROLE_LOCAL_NAME,
    )
    return {
        "uri": facade_uri,
        "class_name": str(class_name or "").strip(),
        "class_uri": str(
            _predicate_by_local_name(
                graph,
                facade,
                IS_FACADE_OF_LOCAL_NAME,
            )
            or ""
        ).strip(),
        "role": str(role or "").strip(),
        "fields": fields,
    }


def _resource_facade_uri(
    package: Mapping[str, Any],
    resource: Mapping[str, Any],
) -> str:
    direct = str(resource.get("facadeUri") or "").strip()
    data_entity = resource.get("dataEntity")
    nested = (
        str(data_entity.get("facade") or "").strip()
        if isinstance(data_entity, dict)
        else ""
    )
    package_entity = package.get("dataEntity")
    package_facade = (
        str(package_entity.get("facade") or "").strip()
        if isinstance(package_entity, dict)
        else ""
    )
    facade_uri = direct or nested or package_facade
    if not facade_uri:
        raise ValueError(
            f"Dataset resource does not declare its facade: "
            f"{resource.get('name', '<unnamed>')}"
        )
    return facade_uri


def _entity_metadata(
    package: Mapping[str, Any],
    resource: Mapping[str, Any],
) -> Dict[str, str]:
    package_entity = package.get("dataEntity")
    resource_entity = resource.get("dataEntity")
    package_values = (
        package_entity if isinstance(package_entity, dict) else {}
    )
    resource_values = (
        resource_entity if isinstance(resource_entity, dict) else {}
    )
    return {
        "id": str(
            resource_values.get("id")
            or package_values.get("id")
            or resource.get("entityUri")
            or ""
        ).strip(),
        "identifier": str(
            resource_values.get("identifier")
            or package_values.get("identifier")
            or resource.get("entityIdentifier")
            or ""
        ).strip(),
        "type": str(
            resource_values.get("type")
            or package_values.get("type")
            or ""
        ).strip(),
    }


def _ifc_type_uri(entity_type: str, class_name: str) -> str:
    if not class_name:
        return ""
    if "#" in entity_type:
        return entity_type.rsplit("#", 1)[0] + "#" + class_name
    if "/" in entity_type:
        return entity_type.rsplit("/", 1)[0] + "/" + class_name
    return class_name


def _instance_uri(
    resource: Mapping[str, Any],
    entity: Mapping[str, str],
    role: str,
    row_number: int,
    record: Mapping[str, Any],
) -> str:
    template = str(resource.get("instanceUriTemplate") or "").strip()
    if template:
        try:
            return template.format(**record)
        except (KeyError, ValueError) as exc:
            raise ValueError(
                f"Could not apply instanceUriTemplate: {template}"
            ) from exc

    entity_id = entity.get("id") or entity.get("identifier")
    if not entity_id:
        raise ValueError("Dataset data entity does not declare an identifier.")
    global_id = str(record.get("GlobalId") or "").strip()
    if global_id and entity_id.rstrip("/").endswith(global_id):
        return entity_id
    normalized_role = quote(
        role or str(resource.get("name") or "instance"),
        safe="",
    )
    return (
        f"{entity_id.rstrip('/')}/row/{row_number:08d}/"
        f"{normalized_role}"
    )


def _literal(value: Any, datatype: str) -> Dict[str, Any]:
    if datatype.endswith(("#string", "/string")):
        normalized: Any = str(value)
    elif isinstance(value, bool):
        normalized: Any = value
    elif isinstance(value, (int, float)):
        normalized = value
    else:
        normalized = str(value)
    result: Dict[str, Any] = {"@value": normalized}
    if datatype:
        result["@type"] = datatype
    return result


def _resource_records(
    resource_path: Path,
    resource: Mapping[str, Any],
) -> list[Dict[str, Any]]:
    suffix = resource_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise ValueError(
                "The 'openpyxl' package is required to materialize "
                "dataset facade instances."
            ) from exc
        excel = resource.get("excel")
        dialect = resource.get("dialect")
        if (
            not isinstance(excel, dict)
            and isinstance(dialect, dict)
            and isinstance(dialect.get("excel"), dict)
        ):
            excel = dialect["excel"]
        sheet_name = (
            str(excel.get("sheet") or "").strip()
            if isinstance(excel, dict)
            else ""
        )
        if not sheet_name:
            sheet_name = str(resource.get("name") or "").strip()
        workbook = load_workbook(
            resource_path,
            data_only=True,
            read_only=True,
        )
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet not found in dataset resource: {sheet_name}"
                )
            rows = workbook[sheet_name].iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows)]
            return [
                dict(zip(headers, values))
                for values in rows
                if any(
                    value is not None and str(value).strip()
                    for value in values
                )
            ]
        except StopIteration:
            return []
        finally:
            workbook.close()

    if suffix == ".csv":
        with resource_path.open(
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as stream:
            return [dict(record) for record in DictReader(stream)]

    if suffix in {".json", ".jsonld"}:
        value = json.loads(resource_path.read_text(encoding="utf-8"))
        if isinstance(value, dict) and isinstance(value.get("@graph"), list):
            value = value["@graph"]
        if not isinstance(value, list):
            raise ValueError(
                f"Dataset JSON resource must contain a list: {resource_path}"
            )
        return [dict(item) for item in value if isinstance(item, dict)]

    raise ValueError(
        f"Unsupported dataset resource for facade materialization: "
        f"{resource_path}"
    )


def _materialize_dataset_instances(
    container_path: Path,
    dataset_path: Path,
    facade_graph: Graph,
) -> tuple[list[Dict[str, Any]], set[Path]]:
    datapackage_path = (
        dataset_path / METADATA_DIRECTORY / DATAPACKAGE_FILENAME
    )
    if not datapackage_path.is_file():
        raise ValueError(
            "Dataset datapackage not found: "
            f"{datapackage_path.relative_to(container_path)}"
        )
    package = _load_json_object(datapackage_path, "dataset datapackage")
    nodes: list[Dict[str, Any]] = []
    resource_files: set[Path] = set()

    for resource in _resource_descriptors(package):
        facade_uri = _resource_facade_uri(package, resource)
        contract = _facade_contract(facade_graph, facade_uri)
        entity = _entity_metadata(package, resource)
        resource_path = _resource_path(
            dataset_path,
            datapackage_path,
            resource,
        )
        resource_files.add(resource_path)
        records = _resource_records(resource_path, resource)
        declared_fields = contract["fields"]

        for row_number, record in enumerate(records, start=1):
            node: Dict[str, Any] = {
                "@id": _instance_uri(
                    resource,
                    entity,
                    contract["role"],
                    row_number,
                    record,
                ),
                str(DCTERMS.conformsTo): [{"@id": facade_uri}],
            }
            type_uri = (
                _ifc_type_uri(
                    entity["type"],
                    contract["class_name"],
                )
                if contract["class_name"]
                else contract["class_uri"]
            )
            if type_uri:
                node["@type"] = [type_uri]

            for field_name, field in declared_fields.items():
                value = record.get(field_name)
                if value is None or not str(value).strip():
                    continue
                node.setdefault(field["predicate"], []).append(
                    _literal(value, field["datatype"])
                )
            nodes.append(node)

    return nodes, resource_files


def serialize_dataset_facades(container_path: Path) -> Dict[str, Any]:
    facade_files = dataset_facade_files(container_path)
    graph = Graph()
    instance_nodes: list[Dict[str, Any]] = []
    resource_files: set[Path] = set()

    for facade_file in facade_files:
        try:
            graph.parse(facade_file, format="turtle")
        except Exception as exc:
            raise ValueError(
                "Invalid dataset facade: "
                f"{facade_file.relative_to(container_path)}"
            ) from exc
        dataset_path = facade_file.parents[2]
        nodes, files = _materialize_dataset_instances(
            container_path,
            dataset_path,
            graph,
        )
        instance_nodes.extend(nodes)
        resource_files.update(files)

    serialized = to_canonical_graph(graph).serialize(format="json-ld")
    raw_payload = json.loads(str(serialized))
    graph_nodes = (
        raw_payload.get("@graph", [])
        if isinstance(raw_payload, dict)
        else raw_payload
    )
    if not isinstance(graph_nodes, list):
        graph_nodes = [graph_nodes]

    payload = {
        "@graph": _canonicalize([*graph_nodes, *instance_nodes])
    }
    artifact_path = container_path / FACADES_JSONLD_RELATIVE_PATH
    _atomic_write_text(
        artifact_path,
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
        + "\n",
    )

    fingerprint = calculate_facades_fingerprint(container_path)
    _embed_jsonld(
        container_path / INDEX_FILENAME,
        payload,
        fingerprint,
    )
    return {
        "payload": payload,
        "fingerprint": fingerprint,
        "facade_files": [
            path.relative_to(container_path).as_posix()
            for path in facade_files
        ],
        "resource_files": [
            path.relative_to(container_path).as_posix()
            for path in sorted(
                resource_files,
                key=lambda item: item.relative_to(container_path).as_posix(),
            )
        ],
        "instance_count": len(instance_nodes),
        "artifact_path": str(artifact_path),
    }


def is_facades_serialized(context: CliContextPort) -> bool:
    try:
        container_path = resolve_container_path(context)
        dataset_facade_files(container_path)
        fingerprint = calculate_facades_fingerprint(container_path)
        artifact_path = container_path / FACADES_JSONLD_RELATIVE_PATH
        index_path = container_path / INDEX_FILENAME
        if not artifact_path.is_file() or not index_path.is_file():
            return False

        payload = json.loads(artifact_path.read_text(encoding="utf-8"))
        document = index_path.read_text(encoding="utf-8")
        if (
            'id="work-stream-jsonld"' in document
            or "window.infoBimWorkStreamData" in document
        ):
            return False
        embedded = _embedded_jsonld(document)
        return (
            isinstance(payload, dict)
            and embedded is not None
            and embedded["fingerprint"] == fingerprint
            and embedded["payload"] == payload
        )
    except (
        OSError,
        TypeError,
        ValueError,
        json.JSONDecodeError,
    ):
        return False


def _canonicalize(value: Any, preserve_list_order: bool = False) -> Any:
    if isinstance(value, dict):
        return {
            key: _canonicalize(
                value[key],
                preserve_list_order=(key == "@list"),
            )
            for key in sorted(value)
        }
    if isinstance(value, list):
        items = [
            _canonicalize(item, preserve_list_order=False)
            for item in value
        ]
        if preserve_list_order:
            return items
        return sorted(
            items,
            key=lambda item: json.dumps(
                item,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
        )
    return value


def _embed_jsonld(
    index_path: Path,
    payload: Dict[str, Any],
    fingerprint: str,
) -> None:
    if not index_path.is_file():
        raise ValueError(f"Container index not found: {index_path}")

    document = index_path.read_text(encoding="utf-8")
    document = _SCRIPT_PATTERN.sub("", document)
    closing_head = "</head>"
    if closing_head not in document:
        raise ValueError(
            f"Container index does not contain a closing head tag: {index_path}"
        )

    serialized = json.dumps(
        payload,
        ensure_ascii=False,
        indent=2,
        sort_keys=True,
    ).replace("</", "<\\/")
    head_prefix = (
        "" if document.partition(closing_head)[0].endswith("\n") else "\n"
    )
    fragment = (
        head_prefix
        + f'  <script id="{SCRIPT_ID}" type="application/ld+json" '
        f'{_FINGERPRINT_ATTRIBUTE}="{fingerprint}">\n'
        f"{serialized}\n"
        "  </script>\n"
    )
    updated = document.replace(
        closing_head,
        fragment + closing_head,
        1,
    )
    _atomic_write_text(index_path, updated)


def _embedded_jsonld(document: str) -> Dict[str, Any] | None:
    match = _SCRIPT_PATTERN.search(document)
    if match is None:
        return None

    script = match.group(0)
    fingerprint_match = re.search(
        rf'{_FINGERPRINT_ATTRIBUTE}=[\"\']([0-9a-f]{{64}})[\"\']',
        script,
        re.IGNORECASE,
    )
    content_match = re.search(
        r">(?P<content>.*)</script>",
        script,
        re.IGNORECASE | re.DOTALL,
    )
    if fingerprint_match is None or content_match is None:
        return None

    payload = json.loads(
        content_match.group("content").replace("<\\/", "</")
    )
    return {
        "fingerprint": fingerprint_match.group(1).lower(),
        "payload": payload,
    }


def _atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_name(f".{path.name}.tmp")
    temporary_path.write_text(content, encoding="utf-8")
    temporary_path.replace(path)


class FacadesSerializedCapability(TransactionCapability):
    METADATA = CapabilityMetadata(
        id=(
            "org.ontobdc.view.plugin.capability.transformation.target."
            "facades_serialized"
        ),
        version="1.0.0",
        name="Facades Serialized",
        description=(
            "Materialize every dataset resource through its facade, serialize "
            "the facade contracts and resulting instances as JSON-LD, and "
            "embed the consolidated graph in the standalone container view."
        ),
        author=["http://kb.elias.eng.br/nid/elias.ttl#Elias"],
        tags=[
            "view",
            "container",
            "dataset",
            "facade",
            "json-ld",
            "transformation",
        ],
        supported_languages=["en", "pt-br"],
    )

    def label(self, lang: str = "en") -> str:
        return ContainerViewProcessState.FACADES_SERIALIZED.label(lang)

    def description(self, lang: str = "en") -> str:
        return ContainerViewProcessState.FACADES_SERIALIZED.description(lang)

    def execute(self, context: CliContextPort) -> Dict[str, Any]:
        if not HardcodedCapability().is_satisfied(context):
            raise RuntimeError(
                "The hardcoded container view was not generated."
            )

        container_path = resolve_container_path(context)
        result = serialize_dataset_facades(container_path)
        return {
            "resulting_state": (
                ContainerViewProcessState.FACADES_SERIALIZED
            ),
            "container_path": str(container_path),
            "facade_count": len(result["facade_files"]),
            "facade_files": result["facade_files"],
            "instance_count": result["instance_count"],
            "resource_files": result["resource_files"],
            "facades_fingerprint": result["fingerprint"],
            "jsonld_path": result["artifact_path"],
            "index_path": str(container_path / INDEX_FILENAME),
        }

    def is_satisfied(self, context: CliContextPort) -> bool:
        return is_facades_serialized(context)
