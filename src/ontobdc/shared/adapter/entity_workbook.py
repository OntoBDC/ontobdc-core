import json
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Sequence


@dataclass(frozen=True)
class EntityWorkbookField:
    """One workbook column and its Frictionless field type."""

    name: str
    field_type: str = "string"

    def to_descriptor(self) -> Dict[str, str]:
        return {
            "name": self.name,
            "type": self.field_type,
        }


@dataclass(frozen=True)
class EntityWorkbookArtifact:
    """Files and validation generated for an entity workbook."""

    workbook_path: Path
    datapackage_path: Path
    worksheet_name: str
    generated_row_count: int
    validation: Dict[str, Any]


class EntityWorkbookAdapter:
    """Read and generate a Frictionless-described workbook for any entity."""

    def read(
        self,
        *,
        workbook_path: Path,
        worksheet_name: str,
        fields: Sequence[EntityWorkbookField],
    ) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise ValueError(
                "The 'openpyxl' package is required to read an entity workbook."
            ) from exc

        resolved_workbook_path: Path = workbook_path.expanduser().resolve()
        if not resolved_workbook_path.is_file():
            return []

        workbook = load_workbook(resolved_workbook_path, data_only=False)
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet not found in entity workbook: {worksheet_name}"
            )

        worksheet = workbook[worksheet_name]
        expected_headers: List[str] = [field.name for field in fields]
        actual_headers: List[str] = [
            str(cell.value or "").strip()
            for cell in worksheet[1]
        ]
        if actual_headers != expected_headers:
            raise ValueError(
                "Entity workbook headers do not match the declared fields. "
                f"Expected {expected_headers}; found {actual_headers}."
            )

        records: List[Dict[str, Any]] = []
        for row in worksheet.iter_rows(
            min_row=2,
            max_col=len(expected_headers),
            values_only=True,
        ):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            records.append(dict(zip(expected_headers, row)))

        workbook.close()
        return records

    def generate(
        self,
        *,
        output_dir: Path,
        workbook_name: str,
        worksheet_name: str,
        fields: Sequence[EntityWorkbookField],
        records: Sequence[Mapping[str, Any]],
        datapackage_path: Path,
        package_name: str,
        resource_name: str,
        primary_key: Sequence[str] = (),
    ) -> EntityWorkbookArtifact:
        try:
            from frictionless import Resource, Schema, formats
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ModuleNotFoundError as exc:
            raise ValueError(
                "The 'openpyxl' and 'frictionless' packages are required "
                "to generate an entity workbook."
            ) from exc

        self._validate_contract(
            worksheet_name=worksheet_name,
            fields=fields,
            primary_key=primary_key,
        )

        resolved_output_dir: Path = output_dir.expanduser().resolve()
        resolved_output_dir.mkdir(parents=True, exist_ok=True)
        resolved_datapackage_path: Path = datapackage_path.expanduser().resolve()
        resolved_datapackage_path.parent.mkdir(parents=True, exist_ok=True)
        workbook_path: Path = self.resolve_writable_path(
            output_dir=resolved_output_dir,
            workbook_name=workbook_name,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = worksheet_name

        field_names: List[str] = [field.name for field in fields]
        worksheet.append(field_names)
        for header_cell in worksheet[1]:
            header_cell.font = Font(bold=True)

        for record in records:
            worksheet.append(
                [
                    self._to_cell_value(record.get(field_name))
                    for field_name in field_names
                ]
            )

        worksheet.freeze_panes = "A2"
        self._fit_columns(worksheet)
        workbook.save(workbook_path)
        workbook.close()

        workbook_relative_path: str = Path(
            os.path.relpath(
                workbook_path,
                start=resolved_datapackage_path.parent,
            )
        ).as_posix()
        schema_descriptor: Dict[str, Any] = {
            "fields": [field.to_descriptor() for field in fields],
        }
        if primary_key:
            schema_descriptor["primaryKey"] = list(primary_key)

        package_resource = Resource(
            name=resource_name,
            path=workbook_relative_path,
            schema=Schema(schema_descriptor),
            control=formats.ExcelControl(sheet=worksheet_name),
        )
        self._write_datapackage(
            datapackage_path=resolved_datapackage_path,
            package_name=package_name,
            resource_descriptor=package_resource.to_descriptor(),
        )

        validation_resource = Resource(
            name=resource_name,
            path=workbook_path.name,
            basepath=str(workbook_path.parent),
            schema=Schema(schema_descriptor),
            control=formats.ExcelControl(sheet=worksheet_name),
        )
        validation_report = validation_resource.validate()
        validation_payload: Dict[str, Any] = validation_report.to_descriptor()
        if not validation_report.valid:
            raise ValueError(
                "Frictionless validation failed for datapackage: "
                f"{json.dumps(validation_payload, ensure_ascii=False)}"
            )

        return EntityWorkbookArtifact(
            workbook_path=workbook_path,
            datapackage_path=resolved_datapackage_path,
            worksheet_name=worksheet_name,
            generated_row_count=len(records),
            validation=validation_payload,
        )

    def resolve_writable_path(
        self,
        *,
        output_dir: Path,
        workbook_name: str,
    ) -> Path:
        base_path: Path = output_dir / workbook_name
        if not base_path.exists():
            return base_path

        try:
            with base_path.open("ab"):
                return base_path
        except PermissionError:
            pass

        for index in range(1, 1000):
            candidate_path: Path = (
                output_dir / f"{base_path.stem}__{index}{base_path.suffix}"
            )
            if not candidate_path.exists():
                return candidate_path
            try:
                with candidate_path.open("ab"):
                    return candidate_path
            except PermissionError:
                continue

        raise PermissionError(
            f"Could not resolve a writable workbook path under: {output_dir}"
        )

    def _write_datapackage(
        self,
        *,
        datapackage_path: Path,
        package_name: str,
        resource_descriptor: Dict[str, Any],
    ) -> None:
        descriptor: Dict[str, Any] = {}
        if datapackage_path.is_file():
            try:
                existing_descriptor: Any = json.loads(
                    datapackage_path.read_text(encoding="utf-8")
                )
                if isinstance(existing_descriptor, dict):
                    descriptor = existing_descriptor
            except (json.JSONDecodeError, OSError):
                descriptor = {}

        descriptor.setdefault("name", package_name)
        existing_resources: List[Dict[str, Any]] = [
            resource
            for resource in descriptor.get("resources", [])
            if isinstance(resource, dict)
            and resource.get("name") != resource_descriptor.get("name")
        ]
        descriptor["resources"] = [
            *existing_resources,
            resource_descriptor,
        ]
        datapackage_path.write_text(
            json.dumps(
                descriptor,
                ensure_ascii=False,
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )

    def _validate_contract(
        self,
        *,
        worksheet_name: str,
        fields: Sequence[EntityWorkbookField],
        primary_key: Sequence[str],
    ) -> None:
        if not worksheet_name.strip():
            raise ValueError("Worksheet name is required.")
        if not fields:
            raise ValueError("At least one entity workbook field is required.")

        field_names: List[str] = [field.name for field in fields]
        if any(not field_name.strip() for field_name in field_names):
            raise ValueError("Entity workbook field names cannot be empty.")
        if len(field_names) != len(set(field_names)):
            raise ValueError("Entity workbook field names must be unique.")

        unknown_primary_keys: List[str] = [
            key for key in primary_key if key not in field_names
        ]
        if unknown_primary_keys:
            raise ValueError(
                "Unknown primary key fields: "
                + ", ".join(unknown_primary_keys)
            )

    def _to_cell_value(self, value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return value

    def _fit_columns(self, worksheet: Any) -> None:
        for column_cells in worksheet.columns:
            width: int = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(width + 2, 12),
                60,
            )
