import json
from pathlib import Path
from typing import Any, Dict

import pytest
from openpyxl import Workbook, load_workbook
from rdflib import Graph, Literal, Namespace, URIRef
from rdflib.namespace import RDF, RDFS, XSD

from ontobdc.view.plugin.capability.facades_serialized import (
    FACADES_JSONLD_RELATIVE_PATH,
    SCRIPT_ID,
    calculate_facades_fingerprint,
    dataset_facade_files,
    is_facades_serialized,
    serialize_dataset_facades,
)


OBDC = Namespace("http://ontobdc.org/ontology/domain/ns.ttl#")
SCHEMA = Namespace("https://schema.org/")
TEST_FACADE = Namespace("urn:test:facade-contract:")


class FakeContext:
    def __init__(self, **parameters: Any) -> None:
        self._parameters: Dict[str, Any] = dict(parameters)

    def get_parameter_value(self, name: str) -> Any:
        return self._parameters.get(name)

    def set_parameter_value(self, name: str, value: Any) -> None:
        self._parameters[name] = value

    def has_parameter(self, name: str) -> bool:
        return name in self._parameters


def _create_dataset(
    container_path: Path,
    name: str,
    facade_subject: str,
    label: str,
) -> Path:
    dataset_path = container_path / name
    metadata_path = dataset_path / ".__ontobdc__"
    linkset_path = metadata_path / "linkset"
    linkset_path.mkdir(parents=True)

    dataset_graph = Graph()
    dataset_graph.add(
        (
            URIRef(f"urn:test:dataset:{name}"),
            RDF.type,
            OBDC.EntityDataset,
        )
    )
    dataset_graph.serialize(
        destination=metadata_path / "dataset.ttl",
        format="turtle",
    )

    facade_graph = Graph()
    subject = URIRef(facade_subject)
    facade_graph.add((subject, RDF.type, URIRef("urn:test:Facade")))
    facade_graph.add((subject, RDFS.label, Literal(label)))
    field = URIRef(f"{facade_subject}:name-field")
    facade_graph.add((subject, TEST_FACADE.hasFacadeField, field))
    facade_graph.add((subject, TEST_FACADE.ifcClassName, Literal("TestEntity")))
    facade_graph.add((subject, TEST_FACADE.instanceRole, Literal("test")))
    facade_graph.add((field, SCHEMA.identifier, Literal("name")))
    facade_graph.add((field, TEST_FACADE.fieldDatatype, XSD.string))
    facade_graph.add((field, TEST_FACADE.mapsToProperty, SCHEMA.name))
    facade_file = linkset_path / "facade.ttl"
    facade_graph.serialize(destination=facade_file, format="turtle")

    document_path = dataset_path / "payload" / "document"
    document_path.mkdir(parents=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "data"
    worksheet.append(["name"])
    worksheet.append([label])
    workbook.save(document_path / "data.xlsx")
    workbook.close()

    (metadata_path / "datapackage.json").write_text(
        json.dumps(
            {
                "dataEntity": {
                    "id": f"urn:test:entity:{name}",
                    "type": "urn:test:ontology:Container",
                },
                "resources": [
                    {
                        "name": "data",
                        "path": "../payload/document/data.xlsx",
                        "excel": {"sheet": "data"},
                        "dataEntity": {"facade": facade_subject},
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    return facade_file


def _create_index(container_path: Path) -> Path:
    index_path = container_path / "index.html"
    index_path.write_text(
        "<!doctype html><html><head><title>Test</title></head>"
        "<body></body></html>\n",
        encoding="utf-8",
    )
    return index_path


def test_serializes_every_dataset_facade_and_embeds_jsonld(
    tmp_path: Path,
) -> None:
    container_path = tmp_path / "container"
    container_path.mkdir()
    first = _create_dataset(
        container_path,
        ".workstream",
        "urn:test:facade:workstream",
        "WorkStream facade",
    )
    second = _create_dataset(
        container_path,
        "cronograma",
        "urn:test:facade:schedule",
        "Schedule facade",
    )
    index_path = _create_index(container_path)

    result = serialize_dataset_facades(container_path)

    assert dataset_facade_files(container_path) == (first, second)
    assert result["facade_files"] == [
        ".workstream/.__ontobdc__/linkset/facade.ttl",
        "cronograma/.__ontobdc__/linkset/facade.ttl",
    ]
    assert result["fingerprint"] == calculate_facades_fingerprint(
        container_path
    )

    artifact = json.loads(
        (container_path / FACADES_JSONLD_RELATIVE_PATH).read_text(
            encoding="utf-8"
        )
    )
    serialized = json.dumps(artifact, ensure_ascii=False)
    assert "urn:test:facade:workstream" in serialized
    assert "urn:test:facade:schedule" in serialized
    assert "WorkStream facade" in serialized
    assert "Schedule facade" in serialized
    assert result["instance_count"] == 2

    html = index_path.read_text(encoding="utf-8")
    assert html.count(f'id="{SCRIPT_ID}"') == 1
    assert 'type="application/ld+json"' in html

    context = FakeContext(container_path=str(container_path))
    assert is_facades_serialized(context)


def test_serialization_is_idempotent(tmp_path: Path) -> None:
    container_path = tmp_path / "container"
    container_path.mkdir()
    _create_dataset(
        container_path,
        "dataset",
        "urn:test:facade",
        "Facade",
    )
    index_path = _create_index(container_path)

    serialize_dataset_facades(container_path)
    first = index_path.read_text(encoding="utf-8")
    serialize_dataset_facades(container_path)
    second = index_path.read_text(encoding="utf-8")

    assert second == first
    assert second.count(f'id="{SCRIPT_ID}"') == 1


def test_changed_facade_invalidates_serialized_state(
    tmp_path: Path,
) -> None:
    container_path = tmp_path / "container"
    container_path.mkdir()
    facade_file = _create_dataset(
        container_path,
        "dataset",
        "urn:test:facade",
        "Original",
    )
    _create_index(container_path)
    context = FakeContext(container_path=str(container_path))

    serialize_dataset_facades(container_path)
    assert is_facades_serialized(context)

    facade_file.write_text(
        facade_file.read_text(encoding="utf-8").replace(
            "Original",
            "Changed",
        ),
        encoding="utf-8",
    )
    assert not is_facades_serialized(context)


def test_changed_dataset_resource_invalidates_serialized_state(
    tmp_path: Path,
) -> None:
    container_path = tmp_path / "container"
    container_path.mkdir()
    _create_dataset(
        container_path,
        "dataset",
        "urn:test:facade",
        "Original",
    )
    _create_index(container_path)
    context = FakeContext(container_path=str(container_path))

    serialize_dataset_facades(container_path)
    assert is_facades_serialized(context)

    workbook_path = (
        container_path / "dataset" / "payload" / "document" / "data.xlsx"
    )
    workbook = load_workbook(workbook_path)
    workbook["data"]["A2"] = "Changed"
    workbook.save(workbook_path)
    workbook.close()

    assert not is_facades_serialized(context)


def test_legacy_workstream_jsonld_invalidates_serialized_state(
    tmp_path: Path,
) -> None:
    container_path = tmp_path / "container"
    container_path.mkdir()
    _create_dataset(
        container_path,
        "dataset",
        "urn:test:facade",
        "Instance",
    )
    index_path = _create_index(container_path)
    context = FakeContext(container_path=str(container_path))

    serialize_dataset_facades(container_path)
    document = index_path.read_text(encoding="utf-8").replace(
        "</body>",
        '<script id="work-stream-jsonld" '
        'type="application/ld+json">{}</script></body>',
    )
    index_path.write_text(document, encoding="utf-8")

    assert not is_facades_serialized(context)


def test_dataset_without_facade_fails_explicitly(tmp_path: Path) -> None:
    container_path = tmp_path / "container"
    metadata_path = container_path / "dataset" / ".__ontobdc__"
    metadata_path.mkdir(parents=True)

    graph = Graph()
    graph.add(
        (
            URIRef("urn:test:dataset"),
            RDF.type,
            OBDC.EntityDataset,
        )
    )
    graph.serialize(
        destination=metadata_path / "dataset.ttl",
        format="turtle",
    )
    _create_index(container_path)

    with pytest.raises(ValueError, match="Dataset facade not found"):
        serialize_dataset_facades(container_path)


def test_embedded_jsonld_escapes_script_closing_sequence(
    tmp_path: Path,
) -> None:
    container_path = tmp_path / "container"
    container_path.mkdir()
    _create_dataset(
        container_path,
        "dataset",
        "urn:test:facade",
        "</script><script>alert(1)</script>",
    )
    index_path = _create_index(container_path)

    serialize_dataset_facades(container_path)

    html = index_path.read_text(encoding="utf-8")
    assert "</script><script>alert(1)</script>" not in html
    assert "<\\/script><script>alert(1)<\\/script>" in html
    assert html.count(f'id="{SCRIPT_ID}"') == 1
